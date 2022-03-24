import openpyxl

class XLUtils:
    @staticmethod
    def getRowCount(file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return(sheet.max_row)
    @staticmethod
    def getColumnCount(file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return(sheet.max_column)
    @staticmethod
    def readData(file,sheetName,rownum,columno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rownum, column=columno).value
    @staticmethod
    def writeData(file,sheetName,rownum,columno,data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rownum, column=columno).value = data
        workbook.save(file)